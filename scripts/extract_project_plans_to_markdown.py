from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
import logging
import re
from typing import Iterable

from docx import Document as load_docx
from docx.document import Document as DocxDocument
from docx.oxml.table import CT_Tbl
from docx.oxml.text.paragraph import CT_P
from docx.table import Table
from docx.text.paragraph import Paragraph
from openpyxl import load_workbook
from pdfminer.high_level import extract_text as pdfminer_extract_text
from pypdf import PdfReader


ROOT = Path(__file__).resolve().parents[1]
SOURCE_DIR = ROOT / "已选项目计划书"
OUTPUT_DIR = ROOT / "项目计划书Markdown"

logging.getLogger("pdfminer").setLevel(logging.ERROR)


@dataclass
class SourceSpec:
    file_name: str
    kind: str
    label: str = "主文档"


@dataclass
class ProjectSpec:
    output_name: str
    status: str
    description: str
    sources: list[SourceSpec]
    notes: list[str] = field(default_factory=list)


PROJECTS: list[ProjectSpec] = [
    ProjectSpec(
        output_name="Vibe Coding 实战训练营",
        status="extracted",
        description="已从 PDF 招生/课程方案文档抽取，当前原始资料未拆分为高级班与初级版两个独立计划书。",
        sources=[SourceSpec("VibeCoding_训练营招生文档.md.pdf", "pdf")],
        notes=["如后续网站需要按班型拆分，可基于此 Markdown 再拆成高级班和初级版两个版本。"],
    ),
    ProjectSpec(
        output_name="OpenClaw 实训营",
        status="extracted",
        description="已从 PDF 课程方案文档抽取。",
        sources=[SourceSpec("OpenClaw_AI助手实战训练营_课程方案.pdf", "pdf")],
    ),
    ProjectSpec(
        output_name="情绪早期干预系统",
        status="extracted",
        description="已从 DOCX 项目计划书抽取。",
        sources=[SourceSpec("心觉卫士—基于多模态的情绪识别早期干预系统.docx", "docx")],
    ),
    ProjectSpec(
        output_name="舌像检测",
        status="extracted",
        description="已从 DOCX 项目计划书抽取。",
        sources=[SourceSpec("舌界之瞳——基于YOLO的舌象多特征实时检测系统.docx", "docx")],
    ),
    ProjectSpec(
        output_name="肠道声学信号",
        status="extracted",
        description="已从 DOCX 项目计划书抽取。",
        sources=[SourceSpec("肠音慧眼-基于深度神经网络的胃肠道声学信号分析与可视化系统.docx", "docx")],
    ),
    ProjectSpec(
        output_name="记忆守护者",
        status="extracted",
        description="已从 PDF 项目计划书抽取。",
        sources=[SourceSpec("记忆守护者.pdf", "pdf")],
    ),
    ProjectSpec(
        output_name="帕金森手环",
        status="extracted",
        description="已从 PDF 项目计划书抽取。",
        sources=[SourceSpec("帕金森检测手环.pdf", "pdf")],
    ),
    ProjectSpec(
        output_name="智能行宠",
        status="extracted",
        description="已从 DOCX 项目计划书抽取。",
        sources=[SourceSpec("智能行宠--30课时实践项目计划书.docx", "docx")],
    ),
    ProjectSpec(
        output_name="遮阳云朵",
        status="extracted",
        description="已从 DOCX 项目计划书抽取。",
        sources=[SourceSpec("遮阳云朵 - 30 课时实践项目计划书.docx", "docx")],
    ),
    ProjectSpec(
        output_name="上肢外骨骼",
        status="extracted",
        description="已从 DOCX 项目计划书抽取。",
        sources=[SourceSpec("助力上肢外骨骼系统.docx", "docx")],
    ),
    ProjectSpec(
        output_name="微风发电",
        status="extracted",
        description="已从 DOCX 项目计划书抽取。",
        sources=[SourceSpec("微风发电系统设计.docx", "docx")],
    ),
    ProjectSpec(
        output_name="AI智眼",
        status="extracted",
        description="已从 PDF 项目计划书抽取。",
        sources=[SourceSpec("ai智眼.pdf", "pdf")],
    ),
    ProjectSpec(
        output_name="闪电小智 AI 宠物狗",
        status="extracted",
        description="已从主计划书和补充稿合并抽取。",
        sources=[
            SourceSpec("闪电小智 AI 宠物狗 — 30课时精品全功能课程计划书.docx", "docx", "主计划书"),
            SourceSpec("宠物.docx", "docx", "补充稿"),
        ],
        notes=["`宠物.docx` 与正式计划书内容高度相关，已作为补充来源合并保留。"],
    ),
    ProjectSpec(
        output_name="智能药盒",
        status="extracted",
        description="已从 DOCX 说明文档抽取，当前资料更偏作品说明和操作说明，不是完整商业计划书版式。",
        sources=[SourceSpec("药盒.docx", "docx")],
    ),
    ProjectSpec(
        output_name="智能花盆-课程大纲",
        status="outline_only",
        description="当前仅有课程大纲，未找到完整项目计划书。",
        sources=[SourceSpec("智能花盆课程大纲.xlsx", "xlsx")],
        notes=["如后续补到正式计划书，可直接覆盖替换该 Markdown。"],
    ),
]


MISSING_ITEMS = [
    ("人型机器人", "未在当前文件夹中找到项目计划书或课程大纲。"),
    ("智能花盆", "当前只找到课程大纲，未找到完整项目计划书。"),
]


def normalize_text(text: str) -> str:
    text = text.replace("\u00a0", " ").replace("\ufeff", "")
    text = re.sub(r"[ \t]+", " ", text)
    return text.strip()


def collapse_blank_lines(lines: Iterable[str]) -> list[str]:
    collapsed: list[str] = []
    blank = False
    for raw_line in lines:
        line = raw_line.rstrip()
        if line:
            collapsed.append(line)
            blank = False
        elif not blank:
            collapsed.append("")
            blank = True
    while collapsed and not collapsed[0]:
        collapsed.pop(0)
    while collapsed and not collapsed[-1]:
        collapsed.pop()
    return collapsed


def iter_block_items(parent: DocxDocument):
    parent_elm = parent.element.body
    for child in parent_elm.iterchildren():
        if isinstance(child, CT_P):
            yield Paragraph(child, parent)
        elif isinstance(child, CT_Tbl):
            yield Table(child, parent)


def paragraph_to_markdown(paragraph: Paragraph) -> list[str]:
    text = normalize_text(paragraph.text)
    if not text:
        return []

    style_name = ""
    if paragraph.style is not None and paragraph.style.name:
        style_name = paragraph.style.name.lower()

    if "title" in style_name:
        return [f"# {text}", ""]

    heading_match = re.search(r"(heading|标题)\s*(\d+)", style_name)
    if heading_match:
        level = max(2, min(6, int(heading_match.group(2)) + 1))
        return [f"{'#' * level} {text}", ""]

    if text.startswith(("•", "●", "·", "▪", "◦")):
        return [f"- {text[1:].strip()}"]

    return [text, ""]


def clean_table_rows(table: Table) -> list[list[str]]:
    rows: list[list[str]] = []
    for row in table.rows:
        cells = [normalize_text(cell.text) for cell in row.cells]
        trimmed = list(cells)
        while trimmed and trimmed[-1] == "":
            trimmed.pop()
        if any(trimmed):
            rows.append(trimmed)
    return rows


def rows_to_markdown(rows: list[list[str]]) -> list[str]:
    if not rows:
        return []

    width = max(len(row) for row in rows)
    normalized_rows = [row + [""] * (width - len(row)) for row in rows]

    if len(normalized_rows) == 1:
        return [f"- {' | '.join(cell or ' ' for cell in normalized_rows[0])}", ""]

    header = normalized_rows[0]
    divider = ["---"] * width
    lines = [
        "| " + " | ".join(cell or " " for cell in header) + " |",
        "| " + " | ".join(divider) + " |",
    ]
    for row in normalized_rows[1:]:
        lines.append("| " + " | ".join(cell or " " for cell in row) + " |")
    lines.append("")
    return lines


def extract_docx(path: Path) -> list[str]:
    document = load_docx(path)
    lines: list[str] = []
    for block in iter_block_items(document):
        if isinstance(block, Paragraph):
            lines.extend(paragraph_to_markdown(block))
        elif isinstance(block, Table):
            rows = clean_table_rows(block)
            lines.extend(rows_to_markdown(rows))
    return collapse_blank_lines(lines)


def extract_pdf(path: Path) -> list[str]:
    try:
        return extract_pdf_with_pypdf(path)
    except Exception:
        return extract_pdf_with_pdfminer(path)


def extract_pdf_with_pypdf(path: Path) -> list[str]:
    reader = PdfReader(str(path))
    lines: list[str] = []
    for index, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        page_lines = format_pdf_text_block(text)
        if not page_lines:
            continue
        lines.append(f"#### 第{index}页")
        lines.append("")
        lines.extend(page_lines)
        lines.append("")
    return collapse_blank_lines(lines)


def extract_pdf_with_pdfminer(path: Path) -> list[str]:
    text = pdfminer_extract_text(str(path)) or ""
    page_texts = [block for block in text.split("\f") if block.strip()]
    lines: list[str] = []
    for index, page_text in enumerate(page_texts, start=1):
        page_lines = format_pdf_text_block(page_text)
        if not page_lines:
            continue
        lines.append(f"#### 第{index}页")
        lines.append("")
        lines.extend(page_lines)
        lines.append("")
    return collapse_blank_lines(lines)


def format_pdf_text_block(text: str) -> list[str]:
    lines: list[str] = []
    for raw_line in text.splitlines():
        line = normalize_text(raw_line)
        if line:
            lines.append(line)
        else:
            lines.append("")
    return collapse_blank_lines(lines)


def extract_xlsx(path: Path) -> list[str]:
    workbook = load_workbook(path, data_only=True)
    lines: list[str] = []
    for worksheet in workbook.worksheets:
        if worksheet.title.lower().startswith("docer_"):
            continue

        raw_rows: list[list[str]] = []
        for row in worksheet.iter_rows(values_only=True):
            cells = [normalize_text("" if value is None else str(value)) for value in row]
            while cells and cells[-1] == "":
                cells.pop()
            if any(cells):
                raw_rows.append(cells)

        if not raw_rows:
            continue

        lines.append(f"#### 工作表：{worksheet.title}")
        lines.append("")

        if raw_rows and sum(1 for cell in raw_rows[0] if cell) == 1:
            lines.append(raw_rows[0][0])
            lines.append("")
            raw_rows = raw_rows[1:]

        if raw_rows:
            lines.extend(rows_to_markdown(raw_rows))

    return collapse_blank_lines(lines)


def extract_source(source: SourceSpec) -> list[str]:
    path = SOURCE_DIR / source.file_name
    if not path.exists():
        raise FileNotFoundError(f"未找到源文件: {path}")

    if source.kind == "docx":
        return extract_docx(path)
    if source.kind == "pdf":
        return extract_pdf(path)
    if source.kind == "xlsx":
        return extract_xlsx(path)
    raise ValueError(f"不支持的文件类型: {source.kind}")


def yaml_quote(value: str) -> str:
    return '"' + value.replace("\\", "\\\\").replace('"', '\\"') + '"'


def build_markdown(project: ProjectSpec) -> str:
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    lines = [
        "---",
        f"title: {yaml_quote(project.output_name)}",
        f"status: {yaml_quote(project.status)}",
        f"description: {yaml_quote(project.description)}",
        f"generated_at: {yaml_quote(generated_at)}",
        "source_files:",
    ]
    for source in project.sources:
        source_path = str((SOURCE_DIR / source.file_name).resolve())
        lines.append(f"  - {yaml_quote(source_path)}")
    lines.extend(["---", "", f"# {project.output_name}", ""])

    lines.extend(
        [
            "## 文档说明",
            "",
            project.description,
            "",
            "## 来源文件",
            "",
        ]
    )

    for source in project.sources:
        lines.append(f"- {source.label}：`{source.file_name}`")

    if project.notes:
        lines.extend(["", "## 备注", ""])
        for note in project.notes:
            lines.append(f"- {note}")

    lines.extend(["", "## 抽取内容", ""])

    for source in project.sources:
        lines.append(f"### {source.label}：{source.file_name}")
        lines.append("")
        lines.extend(extract_source(source))
        lines.append("")

    return "\n".join(collapse_blank_lines(lines)) + "\n"


def build_index(projects: list[ProjectSpec]) -> str:
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    lines = [
        "# 项目计划书 Markdown 索引",
        "",
        f"- 生成时间：{generated_at}",
        f"- 来源目录：`{SOURCE_DIR}`",
        f"- 输出目录：`{OUTPUT_DIR}`",
        "",
        "## 已抽取文档",
        "",
        "| 项目 | 状态 | Markdown 文件 | 来源文件 |",
        "| --- | --- | --- | --- |",
    ]

    for project in projects:
        file_name = f"{project.output_name}.md"
        source_names = "；".join(source.file_name for source in project.sources)
        status_text = "已抽取" if project.status == "extracted" else "仅课程大纲"
        lines.append(f"| {project.output_name} | {status_text} | `{file_name}` | {source_names} |")

    lines.extend(["", "## 缺失与边界项", ""])
    for name, note in MISSING_ITEMS:
        lines.append(f"- `{name}`：{note}")

    lines.extend(
        [
            "",
            "## 使用建议",
            "",
            "- 网站详情页、子链接文案、标签整理，优先基于本目录中的 Markdown 继续结构化处理。",
            "- `Vibe Coding 实战训练营.md` 当前仍是合并资料，若后续要严格区分高级班/初级版，建议在此基础上再拆分。",
            "- `智能花盆-课程大纲.md` 仅能作为课程安排参考，不能替代正式项目计划书。",
        ]
    )

    return "\n".join(lines) + "\n"


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    for project in PROJECTS:
        output_path = OUTPUT_DIR / f"{project.output_name}.md"
        output_path.write_text(build_markdown(project), encoding="utf-8")

    index_path = OUTPUT_DIR / "README.md"
    index_path.write_text(build_index(PROJECTS), encoding="utf-8")

    print(f"已生成 {len(PROJECTS)} 个 Markdown 文件和 1 个索引文件。")
    print(OUTPUT_DIR)


if __name__ == "__main__":
    main()
